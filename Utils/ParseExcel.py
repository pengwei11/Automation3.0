#!/usr/bin/env python
# encoding: utf-8
'''
@author: caopeng
@license: (C) Copyright 2013-2017, Node Supply Chain Manager Corporation Limited.
@contact: 1249294960@qq.com
@software: pengwei
@file: ParseExcel.py
@time: 2019/11/7 13:53
@desc:
'''

from openpyxl import load_workbook
from datetime import datetime
from Utils.ConfigRead import *
from Utils.Logger import Logger
from Utils.ParseYaml import ParseYaml
import xlrd, time
from openpyxl.styles import Border, Side, colors


class ParseExcel(object):
    '''
    解析EXCEL文档
    '''
    def __init__(self, filename):
        self.filename = filename
        self.parseyaml = ParseYaml()
        # 读取excel文件
        self.wb = load_workbook(self.filename)
        # 设置excel单元格边框格式
        self.border_set = Border(left=Side(style='thin', color=colors.BLACK),
                            right=Side(style='thin', color=colors.BLACK),
                            top=Side(style='thin', color=colors.BLACK),
                            bottom=Side(style='thin', color=colors.BLACK))
        self.logger = Logger('logger', LOGS_PATH).getlog()

    def getRowValue(self, sheetname, rowno):
        """
        获取excel某一行的数据
        :param sheetname:
        :param rowno:
        :return:
        """
        try:
            # 获取sheetname对象
            sheetnames = self.wb[sheetname]
            # 创建集合，将指定行内的数据添加进集合
            rowValueList = []
            # 循环所有列
            for i in range(1, sheetnames.max_column+1):
                # 通过行号与列号获取指定单元格信息，并添加进集合
                value = sheetnames.cell(rowno, i).value
                rowValueList.append(value)
            return rowValueList
        except Exception as e:
            self.logger.info(e)
            self.logger.info('读取失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def getColumnValue(self, sheetname, columnno):
        '''
        获取excel某一列的数据
        :param sheetname:
        :param rowno:
        :return:
        '''
        try:
            sheetnames = self.wb[sheetname]
            columnValueList = []
            for i in range(2, sheetnames.max_row+1):
                value = sheetnames.cell(i, columnno).value
                columnValueList.append(value)
            return columnValueList
        except Exception as e:
            self.logger.info(e)
            self.logger.info('读取失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def getMergeColumnValue(self, sheetname, columnno):
        """
        读取合并单元格的数据
        :param sheetname: 工作表
        :param columnno: 列号
        :return:
        """
        try:
            # 获取数据
            data = xlrd.open_workbook(self.filename)
            # 获取所有sheet名字
            sheet_name = data.sheet_by_name(sheetname)
            # 获取总行数
            nrows = sheet_name.nrows  # 包括标题
            # 获取总列数
            ncols = sheet_name.ncols
            # 计算出合并的单元格有哪些
            colspan = {}
            # 如果sheet是合并的单元格 则获取合并单元格的值，并将第一行的数据赋值给合并单元格中的空值
            if sheet_name.merged_cells:
                for item in sheet_name.merged_cells:
                    for row in range(item[0], item[1]):
                        for col in range(item[2], item[3]):
                            # 合并单元格的首格是有值的，所以在这里进行了去重
                            if (row, col) != (item[0], item[2]):
                                colspan.update({(row, col): (item[0], item[2])})

                col = []
                for i in range(1, nrows):
                    if colspan.get((i, columnno-1)):
                        value = sheet_name.cell_value(*colspan.get((i, columnno-1)))
                        col.append(value)
                    else:
                        col.append(sheet_name.cell_value(i, columnno-1))
                return col
        except Exception as e:
            self.logger.info(e)
            self.logger.info('合并单元格读取错误')
        finally:
            self.wb.close()

    def ismerge(self, sheetname):
        """
        判断'工作表'内是否有合并单元格
        :param sheetname:
        :return:
        """
        sheetnames = self.wb[sheetname]
        merge = sheetnames.merged_cells
        return merge

    def getCellValue(self, sheetname, rowno, columnno):
        """
        获取excel某一单元格的数据
        :param sheetname:
        :param rowno:
        :return:
        """
        try:
            sheetnames = self.wb[sheetname]
            CellValue = sheetnames.cell(rowno, columnno).value
            return CellValue
        except Exception as e:
            self.logger.info(e)
            self.logger.info('读取失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def getCellObject(self, sheetname, rowno, columnno):
        """
        获取excel某一单元格的数据
        :param sheetname:
        :param rowno:
        :return:
        """
        try:
            sheetnames = self.wb[sheetname]
            CellValue = sheetnames.cell(rowno, columnno)
            return CellValue
        except Exception as e:
            self.logger.info(e)
            self.logger.info('读取失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def writeCellValue(self, sheetname, rowno, columnno, value):
        """
        向excel某一单元格写入数据
        :param sheetname:
        :param rowno:
        :return:
        """
        try:
            sheetnames = self.wb[sheetname]
            sheetnames.cell(rowno, columnno, value)
            self.wb.save(self.filename)
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
            raise
        except Exception as e:
            self.logger.info(e)
            self.logger.info('写入失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def writeCellTime(self, sheetname, rowno, columnno):
        '''
        向excel某一单元格写入数据
        :param sheetname:
        :param rowno:
        :return:
        '''
        try:
            sh = self.wb[sheetname]
            Time = datetime.now()
            Time.strftime('%Y:%m:%d %H:%M:%S')
            sh.cell(rowno, columnno, Time)
            self.wb.save(self.filename)
            self.logger.info('%s写入时间成功，写入时间为：%s' % (self.filename, Time))
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
        except Exception as e:
            self.logger.info(e)
            self.logger.info('写入失败，请检查工作表名以及行，列号')
        finally:
            self.wb.close()

    def writeCellValues(self, sheetname, rowno, err_info=None, err_pic=None):
        """
        写入 错误信息 错误图片
        :param sheetname:
        :param rowno:
        :param api-result:
        :param err_info:
        :param err_pic:
        :return:
        """
        try:
            self.writeCellTime(sheetname, rowno, testStep_EndTime)
            if err_info:
                self.writeCellValue(sheetname, rowno, testStep_Error, err_info)
            elif err_pic:
                self.writeCellValue(sheetname, rowno, testStep_Picture, err_pic)
            else:
                self.writeCellValue(sheetname, rowno, testStep_Error, '')
                self.writeCellValue(sheetname, rowno, testStep_Picture, '')
            self.wb.save(self.filename)
            self.logger.info('用例测试结果，错误信息，错误图片写入成功')
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
        except Exception as e:
            self.wb.close()
            self.logger.info(e)
            self.logger.info('用例测试结果，错误信息，错误图片写入失败')
        finally:
            self.wb.close()

    def clearCellValue(self, sheetname, rowno):
        """
        清空EXCEL单元格数据
        :param sheetname:
        :param rowno:
        :return:
        """
        try:
            endtime = self.getCellValue(sheetname, rowno, testStep_EndTime)
            result = self.getCellValue(sheetname, rowno, testStep_Result)
            err_info = self.getCellValue(sheetname, rowno, testStep_Error)
            err_pic = self.getCellValue(sheetname, rowno, testStep_Picture)
            if endtime is not None or endtime != '' and result is not None or result != '' and\
                    err_info is not None or err_info != '' and err_pic is not None or err_pic != '':
                self.writeCellValue(sheetname, rowno, testStep_EndTime, '')
                self.writeCellValue(sheetname, rowno, testStep_Result, '')
                self.writeCellValue(sheetname, rowno, testStep_Error, '')
                self.writeCellValue(sheetname, rowno, testStep_Picture, '')
            self.wb.save(self.filename)
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
        except Exception as e:
            self.logger.info(e)
            self.logger.info('数据清空失败')
        finally:
            self.wb.close()

    def clearStepColumnValue(self, sheetname):
        """
        清除执行时间，错误结果，错误信息，错误截图信息
        :param sheetname:
        :param columno:
        :return:
        """
        try:
            self.logger.info('清除"%s"工作表测试结果中，请稍等...' % sheetname)
            endtimes = self.getColumnValue(sheetname, testStep_EndTime)
            err_infos = self.getColumnValue(sheetname, testStep_Error)
            err_pics = self.getColumnValue(sheetname, testStep_Picture)
            for a, b in enumerate(endtimes):
                if b == '测试执行时间':
                    continue
                elif b != '' or b is not None:
                    self.wb[sheetname].cell(a + 2, testStep_EndTime, '')
            for e, f in enumerate(err_infos):
                if f == '错误信息':
                    continue
                elif f != '' or f is not None:
                    self.wb[sheetname].cell(e + 2, testStep_Error, '')
            for g, h in enumerate(err_pics):
                if h == '测试截图':
                    continue
                elif h != '' or h is not None:
                    self.wb[sheetname].cell(g + 2, testStep_Picture, '')
            # 清除用例的测试结果
            for l in range(5):
                results = self.getColumnValue(sheetname, testStep_Result+l)
                NoneResults = list(filter(None, self.getColumnValue(sheetname, testStep_Result+l)))
                if len(NoneResults) == 0:
                    continue
                else:
                    for c, d in enumerate(results):
                        if '测试结果1' == d or '测试结果2' == d or '测试结果3' == d or '测试结果4' == d or '测试结果5' == d:
                            continue
                        elif type(self.getCellObject(sheetname, c + 2, testStep_Result + l)).__name__ == 'MergedCell':
                            continue
                        elif d != '' or d is not None:
                            self.wb[sheetname].cell(c+2, testStep_Result+l, '')
            self.wb.save(self.filename)
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
        except Exception as e:
            self.logger.info(e)
            self.logger.info('数据清空失败')
        finally:
            self.wb.close()


    def clearCaseColumnValue(self, sheetname):
        """
        清除执行时间，错误结果，错误信息，错误截图信息
        :param sheetname:
        :param columno:
        :return:
        """
        try:
            self.logger.info('清除"%s"工作表测试结果中，请稍等....' % sheetname)
            # 清除用例的测试结果
            for l in range(5):
                caseResult = self.getColumnValue(sheetname, testCase_Result+l)
                NoneCaseResult = list(filter(None, self.getColumnValue(sheetname, testCase_Result+l)))
                if len(NoneCaseResult) == 1:
                    continue
                else:
                    for i, v in enumerate(caseResult):
                        if '执行结果1' == v or '执行结果2' == v or '执行结果3' == v or '执行结果4' == v or '执行结果5' == v:
                            continue
                        elif type(self.getCellObject(sheetname, i+2, testCase_Result+l)).__name__ == 'MergedCell':
                            continue
                        elif v != '' or v is not None:
                            self.wb[sheetname].cell(i+2, testCase_Result+l, '')
            # 清除执行时间
            caseTime = self.getColumnValue(sheetname, testCase_EndTime)
            for s, d in enumerate(caseTime):
                if d == '执行结束时间':
                    continue
                elif type(self.getCellObject(sheetname, s + 2, testCase_EndTime)).__name__ == 'MergedCell':
                    continue
                elif d != '' or d is not None:
                    self.wb[sheetname].cell(s+2, testCase_EndTime, '')
            self.wb.save(self.filename)
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
        except Exception as e:
            self.logger.info(e)
            self.logger.info('数据清空失败')
        finally:
            self.wb.close()


    def borderAround(self, sheetname, start_row, start_col):
        """
        start_row:拆分单元格开始行
        start_col:拆分单元格开始列
        start_row:拆分单元格结束行
        start_col:拆分单元格结束列
        """
        try:
            sheetnames = self.wb[sheetname]
            sheetnames.cell(start_row, start_col).border = self.border_set
        except PermissionError:
            self.logger.info('请先关闭用例文件，再运行测试用例')
            raise
        except Exception as e:
            self.logger.info(e)
            self.logger.info('边框修改失败，请检查工作表名以及行，列号')


if __name__ == '__main__':
    print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
    p = ParseExcel(r'E:\Automation3.0\麒麟无纸化测试用例.xlsx')
    print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
    p.writeCellValue('历史会议', 562 ,1, '测试')
    p.writeCellValue('历史会议', 562 ,2, '测试')
    p.writeCellValue('历史会议', 562 ,3, '测试')
    p.writeCellValue('历史会议', 562 ,4, '测试')
    print(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))